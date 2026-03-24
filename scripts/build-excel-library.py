"""
Build the GuitarForge Excel library from codebase data.
Parses spotify-songs.ts, songs-data.ts, and exercises.ts,
then writes a formatted songs-library.xlsx with Songs + Exercises sheets.
"""

import re
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = r"C:\Users\User\guitarforge"
SPOTIFY_FILE = os.path.join(BASE, "src", "lib", "spotify-songs.ts")
SONGS_DATA_FILE = os.path.join(BASE, "src", "lib", "songs-data.ts")
EXERCISES_FILE = os.path.join(BASE, "src", "lib", "exercises.ts")
OUTPUT_FILE = os.path.join(BASE, "songs-library.xlsx")


def parse_ts_objects(text: str) -> list[dict]:
    """Extract JS/TS object literals from an array and parse them."""
    results = []
    # Find all { ... } blocks (non-greedy, handling nested braces minimally)
    # We use a brace-counting approach for robustness
    i = 0
    while i < len(text):
        if text[i] == '{':
            depth = 1
            start = i
            i += 1
            while i < len(text) and depth > 0:
                if text[i] == '{':
                    depth += 1
                elif text[i] == '}':
                    depth -= 1
                i += 1
            obj_str = text[start:i]
            parsed = parse_single_object(obj_str)
            if parsed and 'id' in parsed:
                results.append(parsed)
        else:
            i += 1
    return results


def parse_single_object(obj_str: str) -> dict | None:
    """Parse a single JS object literal string into a dict."""
    result = {}
    # Remove outer braces
    inner = obj_str.strip()
    if inner.startswith('{'):
        inner = inner[1:]
    if inner.endswith('}'):
        inner = inner[:-1]

    # Extract key-value pairs using regex
    # Handle: key: value, key: "string", key: 'string', key: true/false, key: number
    # Also handle template literals with backticks (tex field) and arrays

    pos = 0
    while pos < len(inner):
        # Skip whitespace and commas
        while pos < len(inner) and inner[pos] in ' \t\n\r,':
            pos += 1
        if pos >= len(inner):
            break

        # Match key
        key_match = re.match(r'(\w+)\s*:', inner[pos:])
        if not key_match:
            pos += 1
            continue

        key = key_match.group(1)
        pos += key_match.end()

        # Skip whitespace
        while pos < len(inner) and inner[pos] in ' \t\n\r':
            pos += 1
        if pos >= len(inner):
            break

        # Parse value based on first character
        ch = inner[pos]

        if ch == '"':
            # Double-quoted string
            end = find_closing_quote(inner, pos, '"')
            val = inner[pos + 1:end]
            result[key] = val.replace('\\"', '"').replace('\\n', '\n')
            pos = end + 1

        elif ch == "'":
            # Single-quoted string
            end = find_closing_quote(inner, pos, "'")
            val = inner[pos + 1:end]
            result[key] = val.replace("\\'", "'")
            pos = end + 1

        elif ch == '`':
            # Template literal - skip entirely (tex field)
            end = inner.find('`', pos + 1)
            if end == -1:
                end = len(inner) - 1
            val = inner[pos + 1:end]
            result[key] = val[:100] + "..." if len(val) > 100 else val
            pos = end + 1

        elif ch == '[':
            # Array
            end = inner.find(']', pos)
            if end == -1:
                end = len(inner) - 1
            arr_str = inner[pos + 1:end]
            items = re.findall(r'"([^"]*)"', arr_str)
            result[key] = items
            pos = end + 1

        elif ch in 'tf':
            # Boolean
            if inner[pos:pos + 4] == 'true':
                result[key] = True
                pos += 4
            elif inner[pos:pos + 5] == 'false':
                result[key] = False
                pos += 5
            else:
                pos += 1

        elif ch == '-' or ch.isdigit():
            # Number
            num_match = re.match(r'-?[\d.]+', inner[pos:])
            if num_match:
                val_str = num_match.group(0)
                result[key] = float(val_str) if '.' in val_str else int(val_str)
                pos += num_match.end()
            else:
                pos += 1
        else:
            pos += 1

    return result if result else None


def find_closing_quote(text: str, start: int, quote: str) -> int:
    """Find the closing quote, handling escaped quotes."""
    i = start + 1
    while i < len(text):
        if text[i] == '\\':
            i += 2
            continue
        if text[i] == quote:
            return i
        i += 1
    return len(text) - 1


def load_songs() -> list[dict]:
    """Load and merge songs from both TS files, deduplicating."""
    # Parse manual songs
    with open(SONGS_DATA_FILE, 'r', encoding='utf-8') as f:
        content = f.read()

    # Extract the MANUAL_SONGS array
    manual_start = content.find('MANUAL_SONGS')
    manual_block = content[manual_start:content.find('];', manual_start) + 2]
    manual_songs = parse_ts_objects(manual_block)
    print(f"  Manual songs parsed: {len(manual_songs)}")

    # Parse spotify songs
    with open(SPOTIFY_FILE, 'r', encoding='utf-8') as f:
        spotify_content = f.read()

    spotify_songs = parse_ts_objects(spotify_content)
    print(f"  Spotify songs parsed: {len(spotify_songs)}")

    # Merge with deduplication (manual takes priority)
    seen = set()
    merged = []

    for song in manual_songs:
        key = f"{song.get('title', '').lower()}|{song.get('artist', '').lower()}"
        if key not in seen:
            seen.add(key)
            song['_source'] = 'manual'
            merged.append(song)

    for song in spotify_songs:
        key = f"{song.get('title', '').lower()}|{song.get('artist', '').lower()}"
        if key not in seen:
            seen.add(key)
            song['_source'] = 'spotify'
            merged.append(song)

    # Sort by artist then title
    merged.sort(key=lambda s: (s.get('artist', '').lower(), s.get('title', '').lower()))

    print(f"  Total merged (deduplicated): {len(merged)}")
    return merged


def load_exercises() -> list[dict]:
    """Load exercises from exercises.ts."""
    with open(EXERCISES_FILE, 'r', encoding='utf-8') as f:
        content = f.read()

    exercises = parse_ts_objects(content)
    print(f"  Exercises parsed: {len(exercises)}")
    return exercises


def format_duration(ms) -> str:
    """Convert duration in ms to mm:ss format."""
    if not ms:
        return ""
    total_sec = int(ms) // 1000
    minutes = total_sec // 60
    seconds = total_sec % 60
    return f"{minutes}:{seconds:02d}"


def build_excel(songs: list[dict], exercises: list[dict]):
    """Build the Excel workbook with Songs and Exercises sheets."""
    wb = Workbook()

    # --- Colors and styles ---
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Calibri", size=10)
    cell_align = Alignment(vertical="center", wrap_text=False)
    thin_border = Border(
        left=Side(style="thin", color="333333"),
        right=Side(style="thin", color="333333"),
        top=Side(style="thin", color="333333"),
        bottom=Side(style="thin", color="333333"),
    )

    # ========== SONGS SHEET ==========
    ws_songs = wb.active
    ws_songs.title = "Songs"
    ws_songs.sheet_properties.tabColor = "F59E0B"

    song_headers = [
        "ID", "Title", "Artist", "Album", "Year", "Genre",
        "Difficulty", "Key", "Tempo", "Tuning", "Has GP",
        "GP Path", "Popularity", "Duration", "Artist Country", "Personal"
    ]

    # Write headers
    for col, header in enumerate(song_headers, 1):
        cell = ws_songs.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data
    for row_idx, song in enumerate(songs, 2):
        values = [
            song.get('id', ''),
            song.get('title', ''),
            song.get('artist', ''),
            song.get('album', ''),
            song.get('year', ''),
            song.get('genre', ''),
            song.get('difficulty', ''),
            song.get('key', ''),
            song.get('tempo', ''),
            song.get('tuning', ''),
            'Yes' if song.get('gp') else 'No',
            song.get('gpPath', ''),
            song.get('popularity', ''),
            format_duration(song.get('durationMs')),
            song.get('artistCountry', ''),
            'Yes' if song.get('personal') else '',
        ]
        for col, val in enumerate(values, 1):
            cell = ws_songs.cell(row=row_idx, column=col, value=val)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

    # Column widths
    col_widths = [6, 40, 25, 40, 6, 18, 12, 6, 7, 14, 8, 50, 10, 8, 10, 8]
    for i, w in enumerate(col_widths, 1):
        ws_songs.column_dimensions[get_column_letter(i)].width = w

    # Freeze top row and auto-filter
    ws_songs.freeze_panes = "A2"
    ws_songs.auto_filter.ref = f"A1:{get_column_letter(len(song_headers))}{len(songs) + 1}"

    # ========== EXERCISES SHEET ==========
    ws_ex = wb.create_sheet("Exercises")
    ws_ex.sheet_properties.tabColor = "22C55E"

    ex_headers = [
        "ID", "Name", "Category", "Duration (min)", "BPM Range",
        "Focus Areas", "Description", "Tips", "Has Backing Track",
        "Styles", "GP Path"
    ]

    for col, header in enumerate(ex_headers, 1):
        cell = ws_ex.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, ex in enumerate(exercises, 2):
        styles = ex.get('styles', [])
        styles_str = ', '.join(styles) if isinstance(styles, list) else str(styles) if styles else ''

        values = [
            ex.get('id', ''),
            ex.get('n', ''),
            ex.get('c', ''),
            ex.get('m', ''),
            ex.get('b', ''),
            ex.get('f', ''),
            ex.get('d', ''),
            ex.get('t', ''),
            'Yes' if ex.get('bt') else 'No',
            styles_str,
            ex.get('gpPath', ''),
        ]
        for col, val in enumerate(values, 1):
            cell = ws_ex.cell(row=row_idx, column=col, value=val)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

    ex_col_widths = [6, 50, 14, 12, 12, 35, 60, 60, 14, 25, 50]
    for i, w in enumerate(ex_col_widths, 1):
        ws_ex.column_dimensions[get_column_letter(i)].width = w

    ws_ex.freeze_panes = "A2"
    ws_ex.auto_filter.ref = f"A1:{get_column_letter(len(ex_headers))}{len(exercises) + 1}"

    # ========== SUMMARY SHEET ==========
    ws_sum = wb.create_sheet("Summary", 0)  # Insert at beginning
    ws_sum.sheet_properties.tabColor = "EF4444"

    summary_data = [
        ["GuitarForge Library Summary", ""],
        ["", ""],
        ["Total Songs", len(songs)],
        ["Songs with GP Tabs", sum(1 for s in songs if s.get('gp'))],
        ["Songs without GP Tabs", sum(1 for s in songs if not s.get('gp'))],
        ["Unique Artists", len(set(s.get('artist', '') for s in songs))],
        ["Unique Genres", len(set(s.get('genre', '') for s in songs if s.get('genre')))],
        ["", ""],
        ["Total Exercises", len(exercises)],
        ["Exercise Categories", len(set(e.get('c', '') for e in exercises))],
        ["Exercises with GP", sum(1 for e in exercises if e.get('gpPath'))],
        ["", ""],
        ["Generated", "2026-03-24"],
    ]

    title_font = Font(name="Calibri", size=14, bold=True, color="F59E0B")
    label_font = Font(name="Calibri", size=11, bold=True)
    value_font = Font(name="Calibri", size=11)

    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_a = ws_sum.cell(row=row_idx, column=1, value=label)
        cell_b = ws_sum.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            cell_a.font = title_font
        else:
            cell_a.font = label_font
            cell_b.font = value_font

    ws_sum.column_dimensions['A'].width = 25
    ws_sum.column_dimensions['B'].width = 15

    # Save
    wb.save(OUTPUT_FILE)
    print(f"\nSaved to: {OUTPUT_FILE}")
    print(f"  Songs sheet: {len(songs)} rows")
    print(f"  Exercises sheet: {len(exercises)} rows")


def main():
    print("Building GuitarForge Excel library...")
    print()

    print("Loading songs...")
    songs = load_songs()

    print("\nLoading exercises...")
    exercises = load_exercises()

    print("\nBuilding Excel...")
    build_excel(songs, exercises)

    # Quick stats
    gp_count = sum(1 for s in songs if s.get('gp'))
    artists = len(set(s.get('artist', '') for s in songs))
    genres = set(s.get('genre', '') for s in songs if s.get('genre'))
    print(f"\nStats:")
    print(f"  {len(songs)} songs, {gp_count} with GP tabs, {artists} artists, {len(genres)} genres")
    print(f"  {len(exercises)} exercises across {len(set(e.get('c','') for e in exercises))} categories")
    print("\nDone!")


if __name__ == "__main__":
    main()
