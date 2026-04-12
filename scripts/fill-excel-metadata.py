#!/usr/bin/env python3
"""Fill missing Genre, Difficulty, and Category columns in the GuitarForge Excel library."""

import re
import openpyxl

EXCEL_PATH = r"C:\Users\User\Downloads\GuitarForge Library\GuitarForge_Complete_Library.xlsx"

# ---------------------------------------------------------------------------
# Artist -> Genre mapping (from add-gp-songs.py)
# ---------------------------------------------------------------------------
ARTIST_GENRE = {
    "a perfect circle": "Alternative Metal",
    "acdc": "Hard Rock",
    "afi": "Punk Rock",
    "ace of base": "Pop",
    "adele": "Pop",
    "aerosmith": "Hard Rock",
    "alice cooper": "Hard Rock",
    "alice in chains": "Grunge",
    "alien ant farm": "Alternative Rock",
    "alphaville": "Synth Pop",
    "american hi fi": "Pop Punk",
    "amon amarth": "Melodic Death Metal",
    "andy mckee": "Acoustic",
    "angra": "Power Metal",
    "anouk": "Rock",
    "anthrax": "Thrash Metal",
    "arch enemy": "Melodic Death Metal",
    "arctic monkeys": "Indie Rock",
    "at the gates": "Melodic Death Metal",
    "audioslave": "Alternative Rock",
    "avenged sevenfold": "Heavy Metal",
    "bbking": "Blues",
    "bad religion": "Punk Rock",
    "beach boys": "Classic Rock",
    "beastie boys": "Hip Hop",
    "beatles": "Classic Rock",
    "behemoth": "Death Metal",
    "ben harper": "Blues Rock",
    "black crowes": "Blues Rock",
    "black sabbath": "Heavy Metal",
    "blackberries": "Rock",
    "blind guardian": "Power Metal",
    "blink 182": "Pop Punk",
    "bloodhound gang": "Alternative Rock",
    "blur": "Britpop",
    "bon jovi": "Hard Rock",
    "boston": "Classic Rock",
    "bryan adams": "Rock",
    "buckethead": "Experimental",
    "buddy guy": "Blues",
    "bullet for my valentine": "Metalcore",
    "cannibal corpse": "Death Metal",
    "cash, johnny": "Country",
    "johnny cash": "Country",
    "cat stevens": "Folk",
    "stevens, cat": "Folk",
    "chapman, tracy": "Folk",
    "tracy chapman": "Folk",
    "chet atkins": "Country",
    "chic, the": "Funk",
    "children of bodom": "Melodic Death Metal",
    "chuck berry": "Classic Rock",
    "clapton, eric": "Blues Rock",
    "eric clapton": "Blues Rock",
    "coldplay": "Alternative Rock",
    "cradle of filth": "Black Metal",
    "crazy town": "Rap Rock",
    "creed": "Post-Grunge",
    "crow, sheryl": "Rock",
    "sheryl crow": "Rock",
    "cypress hill": "Hip Hop",
    "daft punk": "Electronic",
    "dark tranquillity": "Melodic Death Metal",
    "dave matthews": "Alternative Rock",
    "david bowie": "Classic Rock",
    "de lucia, paco": "Flamenco",
    "paco de lucia": "Flamenco",
    "dead kennedys": "Punk Rock",
    "death": "Death Metal",
    "deep purple": "Hard Rock",
    "def leppard": "Hard Rock",
    "deftones": "Alternative Metal",
    "dimmu borgir": "Black Metal",
    "dio, ronnie james": "Heavy Metal",
    "ronnie james dio": "Heavy Metal",
    "dire straits": "Classic Rock",
    "disturbed": "Heavy Metal",
    "doctor caspers rabbit show": "Rock",
    "down": "Sludge Metal",
    "dragonforce": "Power Metal",
    "drake, nick": "Folk",
    "nick drake": "Folk",
    "dream theater": "Progressive Metal",
    "drowning pool": "Heavy Metal",
    "dylan, bob": "Folk Rock",
    "bob dylan": "Folk Rock",
    "ed sheeran": "Pop",
    "eminem": "Hip Hop",
    "emmanuel, tommy": "Acoustic",
    "tommy emmanuel": "Acoustic",
    "enigma": "Electronic",
    "eric johnson": "Blues Rock",
    "johnson, eric": "Blues Rock",
    "europe": "Hard Rock",
    "extreme": "Hard Rock",
    "faith no more": "Alternative Metal",
    "fleetwood mac": "Classic Rock",
    "foo fighters": "Alternative Rock",
    "friedman, marty": "Heavy Metal",
    "marty friedman": "Heavy Metal",
    "gallagher, rory": "Blues Rock",
    "rory gallagher": "Blues Rock",
    "gamma ray": "Power Metal",
    "genesis": "Progressive Rock",
    "gilbert, paul": "Hard Rock",
    "paul gilbert": "Hard Rock",
    "gipsy kings": "Flamenco",
    "goldfinger": "Punk Rock",
    "goldman, jean": "Pop",
    "good charlotte": "Pop Punk",
    "gorillaz": "Alternative Rock",
    "grateful dead": "Classic Rock",
    "green day": "Punk Rock",
    "guns n roses": "Hard Rock",
    "guy, buddy": "Blues",
    "harisson, george": "Classic Rock",
    "george harrison": "Classic Rock",
    "heart": "Hard Rock",
    "helloween": "Power Metal",
    "hoobastank": "Post-Grunge",
    "hooker, john lee": "Blues",
    "john lee hooker": "Blues",
    "idol, billy": "Hard Rock",
    "billy idol": "Hard Rock",
    "iggy pop": "Punk Rock",
    "in flames": "Melodic Death Metal",
    "incubus": "Alternative Rock",
    "iron maiden": "Heavy Metal",
    "jackson, michael": "Pop",
    "michael jackson": "Pop",
    "jamiroquai": "Funk",
    "janes addiction": "Alternative Rock",
    "jason mraz": "Pop",
    "jeff buckley": "Alternative Rock",
    "jerry cantrell": "Grunge",
    "jethro tull": "Progressive Rock",
    "jett, joan": "Hard Rock",
    "joan jett": "Hard Rock",
    "jimmy eat world": "Alternative Rock",
    "john mayer": "Blues Rock",
    "mayer, john": "Blues Rock",
    "john, elton": "Classic Rock",
    "elton john": "Classic Rock",
    "johnson, jack": "Acoustic",
    "jack johnson": "Acoustic",
    "joplin, janis": "Blues Rock",
    "janis joplin": "Blues Rock",
    "journey": "Classic Rock",
    "judas priest": "Heavy Metal",
    "kansas": "Progressive Rock",
    "kid rock": "Rock",
    "killers": "Alternative Rock",
    "the killers": "Alternative Rock",
    "killswitch engage": "Metalcore",
    "kings of leon": "Alternative Rock",
    "kiss": "Hard Rock",
    "korn": "Nu Metal",
    "kravitz, lenny": "Rock",
    "lenny kravitz": "Rock",
    "kreator": "Thrash Metal",
    "kyuss": "Stoner Rock",
    "lady gaga": "Pop",
    "lagwagon": "Punk Rock",
    "lamb of god": "Groove Metal",
    "lavigne, avril": "Pop Punk",
    "avril lavigne": "Pop Punk",
    "led zeppelin": "Hard Rock",
    "lennon, john": "Classic Rock",
    "john lennon": "Classic Rock",
    "limp bizkit": "Nu Metal",
    "linkin park": "Nu Metal",
    "living colour": "Hard Rock",
    "lynyrd skynyrd": "Southern Rock",
    "machine head": "Groove Metal",
    "malmsteen, yngwie": "Neoclassical Metal",
    "yngwie malmsteen": "Neoclassical Metal",
    "marilyn manson": "Industrial Metal",
    "maroon 5": "Pop Rock",
    "mccartney, paul": "Classic Rock",
    "paul mccartney": "Classic Rock",
    "mclaughlin, john": "Jazz Fusion",
    "john mclaughlin": "Jazz Fusion",
    "megadeth": "Thrash Metal",
    "meola, al di": "Jazz Fusion",
    "al di meola": "Jazz Fusion",
    "meshuggah": "Progressive Metal",
    "metallica": "Metal",
    "metheny, p": "Jazz",
    "pat metheny": "Jazz",
    "millencolin": "Punk Rock",
    "misfits": "Punk Rock",
    "montgomery, wes": "Jazz",
    "wes montgomery": "Jazz",
    "moore, gary": "Blues Rock",
    "gary moore": "Blues Rock",
    "morbid angel": "Death Metal",
    "morissette, alanis": "Alternative Rock",
    "alanis morissette": "Alternative Rock",
    "motley crue": "Hard Rock",
    "motorhead": "Heavy Metal",
    "muse": "Alternative Rock",
    "mxpx": "Punk Rock",
    "nofx": "Punk Rock",
    "neil young": "Classic Rock",
    "young, neil": "Classic Rock",
    "new found glory": "Pop Punk",
    "nickelback": "Post-Grunge",
    "nightwish": "Symphonic Metal",
    "nine inch nails": "Industrial",
    "nirvana": "Grunge",
    "no doubt": "Ska Punk",
    "no use for a name": "Punk Rock",
    "oasis": "Britpop",
    "offspring": "Punk Rock",
    "opeth": "Progressive Metal",
    "osbourne, ozzy": "Heavy Metal",
    "ozzy osbourne": "Heavy Metal",
    "pantera": "Groove Metal",
    "papa roach": "Nu Metal",
    "paramore": "Pop Rock",
    "pass, joe": "Jazz",
    "joe pass": "Jazz",
    "pearl jam": "Grunge",
    "pennywise": "Punk Rock",
    "perfect": "Rock",
    "pink floyd": "Progressive Rock",
    "pixies": "Alternative Rock",
    "poison": "Hard Rock",
    "presidents of the united states of america": "Alternative Rock",
    "primus": "Alternative Rock",
    "prince": "Pop Rock",
    "puddle of mudd": "Post-Grunge",
    "queen": "Classic Rock",
    "queens of the stone age": "Stoner Rock",
    "r.e.m.": "Alternative Rock",
    "rem": "Alternative Rock",
    "radiohead": "Alternative Rock",
    "rage against the machine": "Rap Metal",
    "rainbow": "Hard Rock",
    "rammstein": "Industrial Metal",
    "ramones": "Punk Rock",
    "the ramones": "Punk Rock",
    "rancid": "Punk Rock",
    "red hot chili peppers": "Alternative Rock",
    "reel big fish": "Ska Punk",
    "reinhardt, django": "Jazz",
    "django reinhardt": "Jazz",
    "renbourn, john": "Folk",
    "john renbourn": "Folk",
    "rise against": "Punk Rock",
    "rush": "Progressive Rock",
    "santana": "Blues Rock",
    "santana, carlos": "Blues Rock",
    "carlos santana": "Blues Rock",
    "satriani, joe": "Hard Rock",
    "joe satriani": "Hard Rock",
    "the satriani, joe": "Hard Rock",
    "scorpions": "Hard Rock",
    "sepultura": "Thrash Metal",
    "sex pistols": "Punk Rock",
    "silverchair": "Alternative Rock",
    "simon, paul": "Folk Rock",
    "paul simon": "Folk Rock",
    "skid row": "Hard Rock",
    "slayer": "Thrash Metal",
    "slipknot": "Nu Metal",
    "smashing pumpkins": "Alternative Rock",
    "sonic magnum": "Rock",
    "soundgarden": "Grunge",
    "springsteen, bruce": "Classic Rock",
    "bruce springsteen": "Classic Rock",
    "staind": "Post-Grunge",
    "steve vai": "Hard Rock",
    "vai, steve": "Hard Rock",
    "stone temple pilots": "Grunge",
    "sublime": "Ska Punk",
    "sum 41": "Pop Punk",
    "symphony x": "Progressive Metal",
    "system of a down": "Alternative Metal",
    "taylor swift": "Pop",
    "taylor, james": "Folk Rock",
    "james taylor": "Folk Rock",
    "temple of the dog": "Grunge",
    "tenacious d": "Comedy Rock",
    "testament": "Thrash Metal",
    "the clash": "Punk Rock",
    "the cure": "Post-Punk",
    "the doors": "Classic Rock",
    "the eagles": "Classic Rock",
    "the hives": "Garage Rock",
    "the james brown band": "Funk",
    "the jews": "Rock",
    "the kinks": "Classic Rock",
    "the police": "New Wave",
    "the prodigy": "Electronic",
    "the rolling stones": "Classic Rock",
    "the smiths": "Indie Rock",
    "the strokes": "Indie Rock",
    "the verve": "Britpop",
    "the vines": "Garage Rock",
    "the who": "Classic Rock",
    "thin lizzy": "Hard Rock",
    "tool": "Progressive Metal",
    "toto": "Classic Rock",
    "townsend, devin": "Progressive Metal",
    "devin townsend": "Progressive Metal",
    "u2": "Rock",
    "ugly kid joe": "Hard Rock",
    "van halen": "Hard Rock",
    "vaughan, stevie ray": "Blues",
    "stevie ray vaughan": "Blues",
    "vicente, amigo": "Flamenco",
    "waters, muddy": "Blues",
    "muddy waters": "Blues",
    "weezer": "Alternative Rock",
    "wheatus": "Pop Rock",
    "whitesnake": "Hard Rock",
    "wonder, stevie": "Soul",
    "stevie wonder": "Soul",
    "zakk wylde & black label society": "Heavy Metal",
    "zakk wylde": "Heavy Metal",
    "black label society": "Heavy Metal",
    "zappa, franck": "Experimental",
    "frank zappa": "Experimental",
    "zz top": "Blues Rock",
}

# Genre -> Difficulty mapping
GENRE_DIFFICULTY = {
    "Thrash Metal": "Advanced",
    "Death Metal": "Advanced",
    "Melodic Death Metal": "Advanced",
    "Black Metal": "Advanced",
    "Progressive Metal": "Advanced",
    "Neoclassical Metal": "Advanced",
    "Power Metal": "Advanced",
    "Jazz Fusion": "Advanced",
    "Jazz": "Advanced",
    "Flamenco": "Advanced",
    "Experimental": "Advanced",
    "Symphonic Metal": "Advanced",
    "Heavy Metal": "Intermediate",
    "Hard Rock": "Intermediate",
    "Alternative Metal": "Intermediate",
    "Groove Metal": "Intermediate",
    "Industrial Metal": "Intermediate",
    "Nu Metal": "Intermediate",
    "Metalcore": "Intermediate",
    "Stoner Rock": "Intermediate",
    "Sludge Metal": "Intermediate",
    "Rap Metal": "Intermediate",
    "Metal": "Intermediate",
    "Classic Rock": "Intermediate",
    "Blues Rock": "Intermediate",
    "Blues": "Intermediate",
    "Progressive Rock": "Intermediate",
    "Alternative Rock": "Intermediate",
    "Grunge": "Intermediate",
    "Southern Rock": "Intermediate",
    "Industrial": "Intermediate",
    "Post-Grunge": "Intermediate",
    "Punk Rock": "Beginner",
    "Pop Punk": "Beginner",
    "Ska Punk": "Beginner",
    "Pop": "Beginner",
    "Pop Rock": "Beginner",
    "Folk": "Beginner",
    "Folk Rock": "Beginner",
    "Acoustic": "Intermediate",
    "Country": "Beginner",
    "Rock": "Intermediate",
    "Britpop": "Intermediate",
    "Indie Rock": "Intermediate",
    "New Wave": "Intermediate",
    "Post-Punk": "Intermediate",
    "Garage Rock": "Intermediate",
    "Synth Pop": "Beginner",
    "Electronic": "Beginner",
    "Hip Hop": "Beginner",
    "Funk": "Intermediate",
    "Soul": "Intermediate",
    "Comedy Rock": "Beginner",
    "Rap Rock": "Intermediate",
}


def normalize_artist(name: str) -> str:
    """Normalize artist name for lookup."""
    n = name.lower().strip()
    if n.startswith("the "):
        n = n[4:]
    if ", " in n:
        parts = n.split(", ", 1)
        n = f"{parts[1]} {parts[0]}"
    n = re.sub(r'([a-z])([A-Z])', r'\1 \2', n)
    n = re.sub(r'[^a-z0-9]', '', n)
    return n


# Build normalized lookup for ARTIST_GENRE
NORM_ARTIST_GENRE = {}
for k, v in ARTIST_GENRE.items():
    norm = re.sub(r'[^a-z0-9]', '', k.lower())
    NORM_ARTIST_GENRE[norm] = v


def lookup_genre(artist_name: str) -> str:
    """Look up genre for an artist, trying multiple normalization strategies."""
    # Strategy 1: direct lowercase match
    low = artist_name.lower().strip()
    if low in ARTIST_GENRE:
        return ARTIST_GENRE[low]

    # Strategy 2: normalized match (strip all special chars)
    norm = normalize_artist(artist_name)
    if norm in NORM_ARTIST_GENRE:
        return NORM_ARTIST_GENRE[norm]

    # Strategy 3: try without "The " prefix on the original
    if low.startswith("the "):
        stripped = low[4:]
        if stripped in ARTIST_GENRE:
            return ARTIST_GENRE[stripped]
        norm2 = re.sub(r'[^a-z0-9]', '', stripped)
        if norm2 in NORM_ARTIST_GENRE:
            return NORM_ARTIST_GENRE[norm2]

    return None


# ---------------------------------------------------------------------------
# Exercise category detection
# ---------------------------------------------------------------------------
CATEGORY_KEYWORDS = [
    # Order matters - more specific patterns first
    (r'\bscale|pentatonic|\bminor\s+scale|\bmajor\s+scale', 'Scales'),
    (r'\bmode|dorian|phrygian|lydian|mixolydian|aeolian|ionian|locrian', 'Modes'),
    (r'\bchord|barre|voicing|power\s*chord|triad', 'Chords'),
    (r'\barpeggio|sweep\s*pick', 'Arpeggios'),
    (r'\blegato|hammer[\s-]*on|pull[\s-]*off', 'Legato'),
    (r'\btap\b|tapping', 'Tapping'),
    (r'\bbend|vibrato', 'Bends'),
    (r'\bshred|speed\s*pick|alternate\s*pick|economy\s*pick|tremolo\s*pick', 'Shred'),
    (r'\brhythm|gallop|palm\s*mut|downpick|chug', 'Rhythm'),
    (r'\bwarm[\s-]*up|stretch|chromatic|finger\s*exercise|spider', 'Warm-Up'),
    (r'\bblues\b', 'Blues'),
    (r'\bjazz\b', 'Jazz'),
    (r'\bfunk\b', 'Funk'),
    (r'\brock\b', 'Rock'),
    (r'\bswing\b', 'Swing'),
    (r'\blick|riff|phrase', 'Licks/Riffs'),
    (r'\bslide\b', 'Slides'),
    (r'\bharmonic', 'Harmonics'),
    (r'\bfinger[\s-]*pick|fingerstyle|travis', 'Fingerpicking'),
    (r'\bclassical\b', 'Classical'),
    (r'\bacoustic\b', 'Acoustic'),
    (r'\bflamenco\b', 'Flamenco'),
    (r'\bfretboard|interval', 'Theory'),
    (r'\btechnique|exercise|etude|study|practice|drill|workout', 'Technique'),
    (r'\bsong|melody|tune|christmas|carol|jingle', 'Songs'),
]


def detect_category(title: str, collection: str) -> str:
    """Detect exercise category from title and collection name."""
    text = f"{collection} {title}".lower()
    for pattern, category in CATEGORY_KEYWORDS:
        if re.search(pattern, text):
            return category
    return "General"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # --- Songs sheet ---
    ws = wb['Songs']
    songs_filled_genre = 0
    songs_filled_diff = 0
    songs_unmatched = set()

    for row in range(2, ws.max_row + 1):
        artist = ws.cell(row, 1).value
        if not artist:
            continue

        # Only fill if empty
        current_genre = ws.cell(row, 6).value
        current_diff = ws.cell(row, 7).value

        genre = lookup_genre(artist)

        if not current_genre:
            if genre:
                ws.cell(row, 6).value = genre
                songs_filled_genre += 1
            else:
                songs_unmatched.add(artist)
                ws.cell(row, 6).value = "Rock"  # fallback
                songs_filled_genre += 1
                genre = "Rock"

        if not current_diff:
            g = genre or (current_genre if current_genre else "Rock")
            diff = GENRE_DIFFICULTY.get(g, "Intermediate")
            ws.cell(row, 7).value = diff
            songs_filled_diff += 1

    # --- Exercises sheet ---
    ws_ex = wb['Exercises']
    ex_filled = 0

    for row in range(2, ws_ex.max_row + 1):
        current_cat = ws_ex.cell(row, 5).value
        if current_cat:
            continue

        collection = ws_ex.cell(row, 1).value or ""
        title = ws_ex.cell(row, 2).value or ""

        category = detect_category(title, collection)
        ws_ex.cell(row, 5).value = category
        ex_filled += 1

    # Save
    wb.save(EXCEL_PATH)

    # Report
    print(f"=== Songs Sheet ===")
    print(f"  Genre filled:      {songs_filled_genre}")
    print(f"  Difficulty filled:  {songs_filled_diff}")
    if songs_unmatched:
        print(f"  Unmatched artists (defaulted to 'Rock'): {len(songs_unmatched)}")
        for a in sorted(songs_unmatched):
            print(f"    - {a}")

    print(f"\n=== Exercises Sheet ===")
    print(f"  Category filled:   {ex_filled}")

    print(f"\nTotal cells filled: {songs_filled_genre + songs_filled_diff + ex_filled}")


if __name__ == "__main__":
    main()
