import openpyxl

wb = openpyxl.load_workbook("C:/Users/User/Downloads/GuitarForge Library/GuitarPro_Library.xlsx")

BASE_EX = "C:\\Users\\User\\Downloads\\GuitarForge Library\\Exercises"
BASE_SONGS = "C:\\Users\\User\\Downloads\\GuitarForge Library\\Songs"

# Exercises sheet
ws = wb["Exercises"]
headers = [cell.value for cell in ws[1]]
print("Exercises headers:", headers)

# Update Source Folder
src_idx = headers.index("Source Folder") if "Source Folder" in headers else None
if src_idx is not None:
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=src_idx + 1, value=BASE_EX)

# Add File Location column
col = len(headers) + 1
ws.cell(row=1, column=col, value="File Location")
for r in range(2, ws.max_row + 1):
    fname = ws.cell(row=r, column=1).value
    if fname:
        ws.cell(row=r, column=col, value=BASE_EX + "\\" + str(fname))

# Songs sheet
ws2 = wb["Songs"]
headers2 = [cell.value for cell in ws2[1]]
print("Songs headers:", headers2)

src_idx2 = headers2.index("Source Folder") if "Source Folder" in headers2 else None
if src_idx2 is not None:
    for r in range(2, ws2.max_row + 1):
        ws2.cell(row=r, column=src_idx2 + 1, value=BASE_SONGS)

col2 = len(headers2) + 1
ws2.cell(row=1, column=col2, value="File Location")
for r in range(2, ws2.max_row + 1):
    fname = ws2.cell(row=r, column=1).value
    if fname:
        ws2.cell(row=r, column=col2, value=BASE_SONGS + "\\" + str(fname))

wb.save("C:/Users/User/Downloads/GuitarForge Library/GuitarPro_Library.xlsx")
print(f"Done! Exercises: {ws.max_row-1} | Songs: {ws2.max_row-1}")
