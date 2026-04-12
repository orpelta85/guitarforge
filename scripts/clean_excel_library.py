# -*- coding: utf-8 -*-
"""
Clean GuitarPro_Library.xlsx:
1. Delete rows marked in red (font color FF9C0006)
2. Translate Hebrew text to English
3. Re-categorize exercises by keywords
4. Delete corresponding .gp/.gpx files for removed rows
"""

import openpyxl
import re
import os
import glob

XLSX_PATH = r'C:\Users\User\Downloads\GuitarForge Library\GuitarPro_Library.xlsx'
EXERCISES_DIR = r'C:\Users\User\Downloads\GuitarForge Library\Exercises'
SONGS_DIR = r'C:\Users\User\Downloads\GuitarForge Library\Songs'

# Hebrew to English translations for exercise/song names
HEBREW_TRANSLATIONS = {
    'דורי': 'Dorian',
    'מודוסים': 'Modes',
    'ללא תפירה': 'seamless',
    'עם תופים': 'with drums',
    'מבוא': 'intro',
    'היפוך': 'inversion',
    'סקסטות': 'Sixths',
    'לכל אקורד בפלייבק בלוז': 'for each chord in blues playback',
    'פנטטוני מורחב פוזיציה ראשונה': 'Extended Pentatonic first position',
    'פנטטוני ראשון': 'Pentatonic first position',
    'פנטטונים': 'Pentatonics',
    'חמשת הפוזיציות': 'five positions',
    'פרזזת עם מתיחה': 'Phrase with stretch',
    'שש עשריות שותקות': 'Sixteenth note rests',
    'דוגמה': 'example',
    'תרגיל ארבעה קולות ללופר': 'Four voice looper exercise',
    'תרגיל מובנה 2 ארבעה קולות ללופר': 'Structured exercise 2 four voice looper',
    'תרגיל מובנה 3 ארבעה קולות ללופר': 'Structured exercise 3 four voice looper',
    'תרגיל מעבר': 'Transition exercise',
    'בפריטה מפורקת': 'in arpeggio picking',
    'שלושת המיתרים': 'three strings',
    # Songs Hebrew
    'טון למטה': 'tone down',
    'גן חיות': 'Zoo',
    'חם ומתוק': 'warm and sweet',
    'ריף': 'riff',
    'פריטה': 'strumming',
    'כיווני פריטה': 'strumming directions',
    'אקורדים': 'chords',
    'שלח לי מלאך': 'Send Me an Angel',
    'פריטה בפזמון': 'strumming in chorus',
    'שלום חנוך': 'Shalom Hanoch',
    'משיח': 'Mashiach',
    'ליין': 'line',
}

# Category keywords mapping
CATEGORY_RULES = [
    (['scale', 'pentatonic', 'mode', 'dorian', 'mixolydian', 'phrygian', 'lydian',
      'aeolian', 'ionian', 'locrian'], 'Modes'),
    (['chord', 'barre', 'power chord', 'voicing', 'triad'], 'Chords'),
    (['arpeggio', 'sweep'], 'Arpeggios/Sweep'),
    (['legato', 'hammer', 'pull'], 'Legato'),
    (['tap', 'tapping'], 'Tapping'),
    (['bend', 'vibrato'], 'Bends'),
    (['shred', 'speed', 'alternate', 'picking'], 'Shred/Picking'),
    (['rhythm', 'gallop', 'palm mute', 'chunk', 'strum'], 'Rhythm'),
    (['warm', 'stretch', 'chromatic'], 'Warm-Up'),
    (['improv', 'jam', 'solo'], 'Improv'),
    (['phrase', 'lick', 'riff'], 'Phrasing/Riffs'),
    (['dynamic', 'accent', 'ghost'], 'Dynamics'),
    (['fretboard', 'caged', 'note', 'position'], 'Fretboard'),
    (['slide'], 'Slide'),
    (['harmonic'], 'Harmonics'),
    (['finger', 'travis', 'hybrid'], 'Picking'),
    (['sequence', 'pattern'], 'Shred/Picking'),  # sequences are typically shred exercises
    (['sixths', 'sixth'], 'Chords'),
    (['looper'], 'Rhythm'),
    (['blues'], 'Phrasing/Riffs'),
]


def is_red_row(row):
    """Check if first cell has red font (FF9C0006 or similar red)."""
    c = row[0]
    if c.font and c.font.color:
        try:
            rgb = c.font.color.rgb
            if isinstance(rgb, str) and '9C0006' in rgb:
                return True
        except (TypeError, ValueError):
            pass
        # Also check for pure red FF0000
        try:
            rgb = c.font.color.rgb
            if isinstance(rgb, str) and len(rgb) == 8:
                r = int(rgb[2:4], 16)
                g = int(rgb[4:6], 16)
                b = int(rgb[6:8], 16)
                if r > 180 and g < 80 and b < 80:
                    return True
        except (TypeError, ValueError):
            pass
    return False


def translate_hebrew(text):
    """Replace Hebrew words/phrases with English translations."""
    if not text or not isinstance(text, str):
        return text
    hebrew_re = re.compile(r'[\u0590-\u05FF]')
    if not hebrew_re.search(text):
        return text

    result = text
    # Sort by length descending so longer phrases match first
    for heb, eng in sorted(HEBREW_TRANSLATIONS.items(), key=lambda x: -len(x[0])):
        result = result.replace(heb, eng)

    # Clean up: remove any remaining Hebrew chars (shouldn't happen but safety)
    result = re.sub(r'[\u0590-\u05FF]+', '', result)
    # Clean up double spaces and leading/trailing
    result = re.sub(r'\s+', ' ', result).strip()
    # Clean up " - -" or "- -" patterns
    result = re.sub(r'\s*-\s*-\s*', ' - ', result)
    result = re.sub(r'^\s*-\s*', '', result)  # remove leading dash
    return result


def categorize_exercise(filename, exercise_name):
    """Categorize based on keywords in filename and exercise name."""
    combined = f"{filename} {exercise_name}".lower()
    for keywords, category in CATEGORY_RULES:
        for kw in keywords:
            if kw in combined:
                return category
    return 'Uncategorized'


def delete_file(filepath):
    """Try to delete a file, print result."""
    if filepath and os.path.exists(filepath):
        try:
            os.remove(filepath)
            print(f"  DELETED file: {filepath}")
            return True
        except Exception as e:
            print(f"  ERROR deleting {filepath}: {e}")
            return False
    else:
        print(f"  File not found (skip): {filepath}")
        return False


def main():
    print("Loading workbook...")
    wb = openpyxl.load_workbook(XLSX_PATH)

    # ========== EXERCISES SHEET ==========
    ws_ex = wb['Exercises']
    print(f"\n=== EXERCISES (rows: {ws_ex.max_row - 1}) ===")

    # Identify red rows and collect file paths to delete
    red_rows_ex = []
    files_to_delete = []

    for row in ws_ex.iter_rows(min_row=2, max_row=ws_ex.max_row):
        if row[0].value is None:
            continue
        if is_red_row(row):
            row_num = row[0].row
            filename = row[0].value
            file_loc = row[5].value if len(row) > 5 else None
            red_rows_ex.append(row_num)
            if file_loc:
                files_to_delete.append(file_loc)
            print(f"  RED (to delete): Row {row_num} - {filename}")

    # Delete red rows (from bottom up to preserve row numbers)
    print(f"\nDeleting {len(red_rows_ex)} red rows from Exercises...")
    for row_num in sorted(red_rows_ex, reverse=True):
        ws_ex.delete_rows(row_num)

    # Now translate Hebrew and re-categorize remaining rows
    print("\nTranslating Hebrew and re-categorizing exercises...")
    translated_count = 0
    recategorized_count = 0

    for row in ws_ex.iter_rows(min_row=2, max_row=ws_ex.max_row):
        if row[0].value is None:
            continue

        # Translate cols: 1 (filename - only exercise name part, keep extension),
        # 2 (exercise name), 6 (file location - DO NOT translate, it's a path)
        # We only translate col 2 (exercise name) since filenames on disk won't change
        old_name = row[1].value
        if old_name and isinstance(old_name, str):
            new_name = translate_hebrew(old_name)
            if new_name != old_name:
                row[1].value = new_name
                translated_count += 1
                print(f"  Translated: '{old_name}' -> '{new_name}'")

        # Re-categorize based on filename + name
        filename = str(row[0].value or '')
        name = str(row[1].value or '')
        old_cat = row[2].value
        new_cat = categorize_exercise(filename, name)
        if new_cat != old_cat:
            row[2].value = new_cat
            recategorized_count += 1
            print(f"  Recategorized: '{name}' : '{old_cat}' -> '{new_cat}'")

    print(f"\nExercises: {translated_count} translations, {recategorized_count} recategorized")

    # ========== SONGS SHEET ==========
    ws_songs = wb['Songs']
    print(f"\n=== SONGS (rows: {ws_songs.max_row - 1}) ===")

    # Check for red rows in songs
    red_rows_songs = []
    for row in ws_songs.iter_rows(min_row=2, max_row=ws_songs.max_row):
        if row[0].value is None:
            continue
        if is_red_row(row):
            row_num = row[0].row
            filename = row[0].value
            file_loc = row[6].value if len(row) > 6 else None
            red_rows_songs.append(row_num)
            if file_loc:
                files_to_delete.append(file_loc)
            print(f"  RED (to delete): Row {row_num} - {filename}")

    for row_num in sorted(red_rows_songs, reverse=True):
        ws_songs.delete_rows(row_num)

    # Translate Hebrew in songs (col 2 = song name)
    song_translated = 0
    for row in ws_songs.iter_rows(min_row=2, max_row=ws_songs.max_row):
        if row[0].value is None:
            continue
        old_name = row[1].value
        if old_name and isinstance(old_name, str):
            new_name = translate_hebrew(old_name)
            if new_name != old_name:
                row[1].value = new_name
                song_translated += 1
                print(f"  Translated song: '{old_name}' -> '{new_name}'")

    print(f"\nSongs: {song_translated} translations")

    # ========== SAVE ==========
    print(f"\nSaving workbook to {XLSX_PATH}...")
    wb.save(XLSX_PATH)
    print("Workbook saved.")

    # ========== DELETE FILES ==========
    print(f"\n=== DELETING {len(files_to_delete)} FILES ===")
    deleted = 0
    for fp in files_to_delete:
        if delete_file(fp):
            deleted += 1

    print(f"\n=== SUMMARY ===")
    print(f"  Exercises red rows deleted: {len(red_rows_ex)}")
    print(f"  Songs red rows deleted: {len(red_rows_songs)}")
    print(f"  Exercises remaining: {ws_ex.max_row - 1}")
    print(f"  Songs remaining: {ws_songs.max_row - 1}")
    print(f"  Hebrew texts translated: {translated_count + song_translated}")
    print(f"  Exercises recategorized: {recategorized_count}")
    print(f"  Files deleted from disk: {deleted}/{len(files_to_delete)}")


if __name__ == '__main__':
    main()
