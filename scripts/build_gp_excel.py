import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load code songs (source of truth for the app)
with open("C:/Users/User/guitarforge/tmp_gf_songs.json", "r", encoding="utf-8") as f:
    code_songs = json.load(f)

# Deduplicate like the code does
seen = set()
unique_songs = []
for s in code_songs:
    key = f"{s['title'].lower()}|{s['artist'].lower()}"
    if key not in seen:
        seen.add(key)
        unique_songs.append(s)

print(f"Unique code songs: {len(unique_songs)}")

# Load Excel GP library for cross-reference
excel_path = "C:/Users/User/Downloads/GuitarForge Library/GuitarForge_Complete_Library.xlsx"

try:
    xf = pd.ExcelFile(excel_path)
    sheet = "Library" if "Library" in xf.sheet_names else xf.sheet_names[0]
    df_excel = pd.read_excel(excel_path, sheet_name=sheet)
except Exception:
    df_excel = pd.read_excel(excel_path)

print(f"Excel GP rows: {len(df_excel)}")
print(f"Excel columns: {list(df_excel.columns)}")

# Build Excel lookup - handle both old and new column names
title_col = "Song Title" if "Song Title" in df_excel.columns else df_excel.columns[1]
artist_col = "Artist" if "Artist" in df_excel.columns else df_excel.columns[0]

excel_lookup = {}
for _, row in df_excel.iterrows():
    title = str(row.get(title_col, "")).strip().lower()
    artist = str(row.get(artist_col, "")).strip().lower()
    key = f"{title}|{artist}"
    if key not in excel_lookup:
        excel_lookup[key] = {
            "file_path": str(row.get("File Path", row.get("GP File (Excel)", ""))),
        }

print(f"Excel lookup entries: {len(excel_lookup)}")

# Build the comprehensive merged library
rows = []
for s in unique_songs:
    key = f"{s['title'].lower()}|{s['artist'].lower()}"
    in_excel = key in excel_lookup
    excel_file = excel_lookup.get(key, {}).get("file_path", "")
    if excel_file == "nan":
        excel_file = ""

    has_gp_code = s["hasGp"]
    gp_path_code = s["gpPath"] if s["gpPath"] else ""
    has_gp_final = "Yes" if (has_gp_code or in_excel) else "No"

    rows.append({
        "Artist": s["artist"],
        "Song Title": s["title"],
        "Has GP Tab": has_gp_final,
        "GP Path (Code)": gp_path_code,
        "In GP Excel": "Yes" if in_excel else "No",
        "GP File (Excel)": excel_file,
        "Has Key": "Yes" if s["key"] else "No",
        "Has Tempo": "Yes" if s["tempo"] else "No",
        "Has Tuning": "Yes" if s["tuning"] else "No",
        "Key": s["key"] if s["key"] else "",
        "Tempo": int(s["tempo"]) if s["tempo"] else "",
        "Tuning": s["tuning"] if s["tuning"] else "",
    })

df = pd.DataFrame(rows)

# Sort: No GP first, then by artist
df["_sort"] = df["Has GP Tab"].map({"No": 0, "Yes": 1})
df = df.sort_values(["_sort", "Artist", "Song Title"]).drop(columns=["_sort"]).reset_index(drop=True)

# Stats
total = len(df)
yes_gp = int((df["Has GP Tab"] == "Yes").sum())
no_gp = int((df["Has GP Tab"] == "No").sum())
pct_gp = round(yes_gp / total * 100, 1)
has_key = int((df["Has Key"] == "Yes").sum())
has_tempo = int((df["Has Tempo"] == "Yes").sum())
has_tuning = int((df["Has Tuning"] == "Yes").sum())
in_excel_count = int((df["In GP Excel"] == "Yes").sum())

print(f"\n=== STATS ===")
print(f"Total songs: {total}")
print(f"With GP: {yes_gp} ({pct_gp}%)")
print(f"Without GP: {no_gp} ({round(100 - pct_gp, 1)}%)")
print(f"With Key: {has_key}")
print(f"With Tempo: {has_tempo}")
print(f"With Tuning: {has_tuning}")
print(f"In GP Excel: {in_excel_count}")

# Save
output_path = excel_path
df.to_excel(output_path, index=False, sheet_name="Library")

# Format with openpyxl
wb = load_workbook(output_path)
ws = wb["Library"]

# Summary sheet
summary = wb.create_sheet("Summary", 0)

summary["A1"] = "GuitarForge Library - GP Tab Status Report"
summary["A1"].font = Font(bold=True, size=14, name="Arial")
summary["A3"] = "Metric"
summary["B3"] = "Value"
summary["A3"].font = Font(bold=True, size=11, name="Arial")
summary["B3"].font = Font(bold=True, size=11, name="Arial")

stats_data = [
    ("Total Songs in Library", total, False),
    ("Songs WITH GP Tab", yes_gp, False),
    ("Songs WITHOUT GP Tab (Priority!)", no_gp, True),
    ("% Coverage (GP Tabs)", f"{pct_gp}%", False),
    ("Songs with Key metadata", has_key, False),
    ("Songs with Tempo metadata", has_tempo, False),
    ("Songs with Tuning metadata", has_tuning, False),
    ("Songs matched in GP Excel file", in_excel_count, False),
    ("Songs NOT in GP Excel", total - in_excel_count, False),
]

for i, (label, val, is_red) in enumerate(stats_data, 4):
    summary[f"A{i}"] = label
    summary[f"B{i}"] = val
    if is_red:
        summary[f"A{i}"].font = Font(bold=True, color="9C0006", name="Arial")
        summary[f"B{i}"].font = Font(bold=True, color="9C0006", name="Arial")
        summary[f"B{i}"].fill = PatternFill("solid", fgColor="FFC7CE")

summary.column_dimensions["A"].width = 40
summary.column_dimensions["B"].width = 15

# Format Library sheet
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
green_fill = PatternFill("solid", fgColor="C6EFCE")
red_fill = PatternFill("solid", fgColor="FFC7CE")
green_font = Font(color="006100", name="Arial", size=10)
red_font = Font(color="9C0006", name="Arial", size=10)
body_font = Font(name="Arial", size=10)
thin_border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

for col_idx in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

widths = [25, 35, 12, 50, 12, 50, 10, 12, 12, 8, 8, 15]
for i, w in enumerate(widths, 1):
    if i <= ws.max_column:
        ws.column_dimensions[get_column_letter(i)].width = w

gp_col = list(df.columns).index("Has GP Tab") + 1
for row_idx in range(2, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = body_font
        cell.border = thin_border

    gp_cell = ws.cell(row=row_idx, column=gp_col)
    if gp_cell.value == "Yes":
        gp_cell.fill = green_fill
        gp_cell.font = green_font
    else:
        gp_cell.fill = red_fill
        gp_cell.font = red_font
    gp_cell.alignment = Alignment(horizontal="center")

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

wb.save(output_path)
print(f"\nSaved to: {output_path}")

# Print first songs WITHOUT GP
no_gp_df = df[df["Has GP Tab"] == "No"].head(15)
print(f"\nFirst 15 songs WITHOUT GP tab:")
for _, r in no_gp_df.iterrows():
    print(f"  {r['Artist']} - {r['Song Title']}")

# Save stats for report
with open("C:/Users/User/guitarforge/tmp_gf_stats.json", "w") as f:
    json.dump({
        "total": total, "yes_gp": yes_gp, "no_gp": no_gp,
        "pct_gp": pct_gp, "has_key": has_key, "has_tempo": has_tempo,
        "has_tuning": has_tuning, "in_excel": in_excel_count,
        "not_in_excel": total - in_excel_count,
    }, f)
