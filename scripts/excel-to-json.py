"""
Convert GuitarProTabs Excel to JSON for the Node.js download script.
Reads Artists + Exercises sheets, outputs gp-artists-list.json
"""
import json
import openpyxl
import os

EXCEL_PATH = r'C:\Users\User\Downloads\GuitarForge Library\All_Artists_GuitarProTabs.xlsx'
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), 'gp-artists-list.json')

wb = openpyxl.load_workbook(EXCEL_PATH)

# Sheet 1: Artists
ws = wb['Artists']
artists = []
for row in ws.iter_rows(min_row=2, values_only=True):
    clean_name = row[0]
    if not clean_name:
        continue
    artists.append({
        'cleanName': str(clean_name).strip(),
        'originalName': str(row[1] or clean_name).strip(),
        'numTabs': int(row[2] or 0),
        'url': str(row[7] or '').strip(),
        'type': 'artist'
    })

# Sheet 2: Exercises & Other
ws2 = wb['Exercises & Other']
exercises = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    name = row[0]
    if not name:
        continue
    # Some rows at the bottom have Artists-sheet format (name, name, numTabs, popularity, ...)
    # Detect by checking if row[2] is a URL string
    url_val = row[2]
    if isinstance(url_val, str) and url_val.startswith('http'):
        # Standard format: Name, NumTabs, URL, Category, Description
        exercises.append({
            'cleanName': str(name).strip(),
            'originalName': str(name).strip(),
            'numTabs': int(row[1] or 0),
            'url': url_val.strip(),
            'type': 'exercise',
            'category': str(row[3] or '').strip(),
        })
    elif isinstance(url_val, int):
        # Artist-format row: Name, OrigName, NumTabs, Popularity, ...
        # Skip these or treat as artists without URL
        print(f'  Skipping misformatted row: {name} (no URL)')
    else:
        print(f'  Skipping row with no URL: {name}')

data = {
    'artists': artists,
    'exercises': exercises,
    'totalArtists': len(artists),
    'totalExercises': len(exercises),
}

with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    json.dump(data, f, indent=2, ensure_ascii=False)

print(f'Exported {len(artists)} artists + {len(exercises)} exercises to {OUTPUT_PATH}')
